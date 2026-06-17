"""Excel 生成器：按模板格式输出"""
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter


HEADER_FILL = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
HEADER_FONT = Font(name='微软雅黑', bold=True, size=10, color='FFFFFF')
TOTAL_FONT = Font(name='微软雅黑', bold=True, size=10)
NORMAL_FONT = Font(name='微软雅黑', size=10)
PLATE_FONT = Font(name='微软雅黑', bold=True, size=10, color='8B0000')
COMPETITOR_FONT = Font(name='微软雅黑', bold=True, size=10)
MONTHLY_TOTAL_FONT = Font(name='微软雅黑', bold=True, size=10, color='8B0000')
THIN_BORDER = Border(
    left=Side(style='thin'), right=Side(style='thin'),
    top=Side(style='thin'), bottom=Side(style='thin')
)
CENTER_WRAP = Alignment(horizontal='center', vertical='center', wrap_text=True)

# 非月份列（需要每户型 2 行合并）
NON_MONTH_COLS = {'板块', '竞品', '业态', '编码', '户型', '户型面积', '建面分段',
                  '整盘套数', '户配', '供应套数', '成交套数', '成交均价',
                  '整盘去化率', '已供去化率', '已供去化占比',
                  '已取证库存', '整盘库存', '近3月月均销量', '近3月月均销量占比'}
PCT_COLS = {'户配', '整盘去化率', '已供去化率', '已供去化占比', '近3月月均销量占比'}


def _make_base_cols():
    return ['板块', '竞品', '业态', '编码', '户型', '户型面积', '建面分段',
            '整盘套数', '户配', '供应套数', '成交套数', '成交均价',
            '整盘去化率', '已供去化率', '已供去化占比']


def _make_trailing_cols():
    return ['近3月月均销量', '近3月月均销量占比', '已取证库存', '整盘库存']


def _make_month_cols(month_labels):
    """每月 1 列，数据分两行（上=套数，下=均价）"""
    return list(month_labels)


def _write_cell(ws, row, col, value, font=None, fmt=None, merge_info=None):
    """写单元格，统一字体边框"""
    cell = ws.cell(row=row, column=col)
    cell.value = value
    cell.font = font or NORMAL_FONT
    cell.border = THIN_BORDER
    cell.alignment = CENTER_WRAP
    if fmt:
        cell.number_format = fmt
    return cell


def _write_data_row(ws, excel_row, all_cols, row_data, month_labels, is_price_row=False):
    """写一行数据"""
    for col_idx, col_name in enumerate(all_cols, 1):
        if col_name in NON_MONTH_COLS:
            if is_price_row:
                continue
            df_col = '面积范围' if col_name == '户型面积' else col_name
            value = row_data.get(df_col, row_data.get(col_name, ''))
        else:
            # 月份列：上行取套数，下行取均价
            suffix = '_均价' if is_price_row else '_套数'
            value = row_data.get(f"{col_name}{suffix}", '')

        if value is None:
            value = ''

        fmt = None
        if col_name in PCT_COLS and isinstance(value, (int, float)) and value != '':
            fmt = '0%'
        if '均价' in col_name and isinstance(value, (int, float)):
            fmt = '#,##0'
        if col_name == '近3月月均销量' and isinstance(value, (int, float)):
            fmt = '0'

        _write_cell(ws, excel_row, col_idx, value, fmt=fmt)


def generate_excel(
    df: pd.DataFrame,
    total_supply: int,
    total_trans: int,
    project_total: int,
    month_labels: list[str],
    project_name: str,
    output_path: str,
    plate: str = '',
    competitor: str = '',
    open_date: str = '',
    plot_ratio: str = '',
    total_avg_price: str = '',
    monthly_total_prices: dict = None,
) -> str:
    """生成户型盘点 Excel（每个户型 2 行）"""
    wb = Workbook()
    ws = wb.active
    ws.title = project_name[:31]

    if monthly_total_prices is None:
        monthly_total_prices = {}

    base_cols = _make_base_cols()
    month_cols = _make_month_cols(month_labels)
    trailing_cols = _make_trailing_cols()
    all_cols = base_cols + month_cols + trailing_cols

    # === 表头 ===
    for col_idx, col_name in enumerate(all_cols, 1):
        _write_cell(ws, 1, col_idx, col_name, font=HEADER_FONT)
        ws.cell(row=1, column=col_idx).fill = HEADER_FILL

    # === 数据行（每户型 2 行）===
    excel_row = 2
    for _, row in df.iterrows():
        # 上行：套数行（全部列有数据）
        _write_data_row(ws, excel_row, all_cols, row, month_labels, is_price_row=False)
        excel_row += 1
        # 下行：均价行（仅月份列有数据）
        _write_data_row(ws, excel_row, all_cols, row, month_labels, is_price_row=True)

        # 合并非月份列的 2 行
        for col_idx, col_name in enumerate(all_cols, 1):
            if col_name in NON_MONTH_COLS:
                ws.merge_cells(start_row=excel_row - 1, start_column=col_idx,
                               end_row=excel_row, end_column=col_idx)

        excel_row += 1

    # === 合计行（2 行）===
    total_row_start = excel_row
    total_units_col = {}
    total_price_col = {}
    for ml in month_labels:
        total_units_col[ml] = sum(int(row.get(f"{ml}_套数", 0) or 0) for _, row in df.iterrows())
        total_price_col[ml] = monthly_total_prices.get(ml, '')

    # 上行：合计行套数
    for col_idx, col_name in enumerate(all_cols, 1):
        if col_name in NON_MONTH_COLS:
            val_map = {
                '编码': '合计', '户型': '', '户型面积': '',
                '建面分段': '-', '整盘套数': project_total, '户配': 1.0,
                '供应套数': total_supply, '成交套数': total_trans,
                '成交均价': total_avg_price,
                '整盘去化率': total_trans / project_total if project_total > 0 else 0,
                '已供去化率': total_trans / total_supply if total_supply > 0 else 0,
                '已供去化占比': '-', '已取证库存': total_supply - total_trans,
                '整盘库存': project_total - total_trans,
                '近3月月均销量': round(df['近3月月均销量'].sum()) if '近3月月均销量' in df.columns else 0,
                '近3月月均销量占比': '-',
            }
            value = val_map.get(col_name, '')
            fmt = '0%' if col_name in PCT_COLS and isinstance(value, (int, float)) else None
        else:
            value = total_units_col.get(col_name, 0)
            fmt = None
        _write_cell(ws, total_row_start, col_idx, value, font=TOTAL_FONT, fmt=fmt)

    # 下行：合计行均价
    total_row_end = total_row_start + 1
    for col_idx, col_name in enumerate(all_cols, 1):
        if col_name in NON_MONTH_COLS:
            continue
        value = total_price_col.get(col_name, '')
        _write_cell(ws, total_row_end, col_idx, value, font=MONTHLY_TOTAL_FONT)

    # 合计行非月份列 2 行合并
    for col_idx, col_name in enumerate(all_cols, 1):
        if col_name in NON_MONTH_COLS:
            ws.merge_cells(start_row=total_row_start, start_column=col_idx,
                           end_row=total_row_end, end_column=col_idx)

    # 合计行：编码+户型+户型面积三列合并
    code_col = all_cols.index('编码') + 1
    area_col = all_cols.index('户型面积') + 1
    ws.merge_cells(start_row=total_row_start, start_column=code_col,
                   end_row=total_row_end, end_column=area_col)
    mc = ws.cell(row=total_row_start, column=code_col)
    mc.value = '合计'
    mc.font = TOTAL_FONT
    mc.alignment = CENTER_WRAP

    # === 板块列合并 ===
    plate_col = all_cols.index('板块') + 1
    if plate:
        ws.merge_cells(start_row=2, start_column=plate_col, end_row=total_row_end, end_column=plate_col)
        pc = ws.cell(row=2, column=plate_col)
        pc.value = plate
        pc.font = PLATE_FONT

    # === 竞品列合并 ===
    competitor_col = all_cols.index('竞品') + 1
    competitor_content = project_name
    if open_date:
        competitor_content += f"\n{open_date}"
    if plot_ratio:
        competitor_content += f"\n{plot_ratio}"
    ws.merge_cells(start_row=2, start_column=competitor_col, end_row=total_row_end, end_column=competitor_col)
    cc = ws.cell(row=2, column=competitor_col)
    cc.value = competitor_content
    cc.font = COMPETITOR_FONT

    # === 列宽 ===
    base_widths = {
        '板块': 8, '竞品': 10, '业态': 12, '编码': 10, '户型': 10,
        '户型面积': 10, '建面分段': 12, '整盘套数': 10, '户配': 8,
        '供应套数': 10, '成交套数': 10, '成交均价': 10,
        '整盘去化率': 10, '已供去化率': 10, '已供去化占比': 10,
        '已取证库存': 10, '整盘库存': 10,
        '近3月月均销量': 12, '近3月月均销量占比': 12,
    }
    for col_idx, col_name in enumerate(all_cols, 1):
        w = base_widths.get(col_name, 10)
        ws.column_dimensions[get_column_letter(col_idx)].width = w

    ws.freeze_panes = 'A2'
    wb.save(output_path)
    return output_path


def generate_multi_sheet_excel(all_results: list[dict], output_path: str) -> str:
    """批量多项目输出，共用一个 Sheet，按 2 行格式"""
    from openpyxl import Workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "户型盘点"

    if not all_results:
        wb.save(output_path)
        return output_path

    first = all_results[0]
    month_labels = first['month_labels']
    base_cols = _make_base_cols()
    month_cols = _make_month_cols(month_labels)
    trailing_cols = _make_trailing_cols()
    all_cols = base_cols + month_cols + trailing_cols

    # 表头
    for col_idx, col_name in enumerate(all_cols, 1):
        _write_cell(ws, 1, col_idx, col_name, font=HEADER_FONT)
        ws.cell(row=1, column=col_idx).fill = HEADER_FILL

    current_row = 2

    for proj_idx, result in enumerate(all_results):
        df = result['df']
        project_name = result['project_name']
        plate = result.get('plate', '')
        open_date = result.get('open_date', '')
        plot_ratio = result.get('plot_ratio', '')
        total_supply = result['total_supply']
        total_trans = result['total_trans']
        project_total = result['project_total']
        total_avg_price = result.get('total_avg_price', '')
        monthly_total_prices = result.get('monthly_total_prices', {})
        if monthly_total_prices is None:
            monthly_total_prices = {}

        data_start = current_row

        # 数据行
        for _, row in df.iterrows():
            _write_data_row(ws, current_row, all_cols, row, month_labels, is_price_row=False)
            current_row += 1
            _write_data_row(ws, current_row, all_cols, row, month_labels, is_price_row=True)
            for col_idx, col_name in enumerate(all_cols, 1):
                if col_name in NON_MONTH_COLS:
                    ws.merge_cells(start_row=current_row - 1, start_column=col_idx,
                                   end_row=current_row, end_column=col_idx)
            current_row += 1

        # 合计行
        total_row_start = current_row
        total_units_col = {}
        total_price_col = {}
        for ml in month_labels:
            total_units_col[ml] = sum(int(row.get(f"{ml}_套数", 0) or 0) for _, row in df.iterrows())
            total_price_col[ml] = monthly_total_prices.get(ml, '')

        for col_idx, col_name in enumerate(all_cols, 1):
            if col_name in NON_MONTH_COLS:
                val_map = {
                    '编码': '合计', '户型': '', '户型面积': '',
                    '建面分段': '-', '整盘套数': project_total, '户配': 1.0,
                    '供应套数': total_supply, '成交套数': total_trans,
                    '成交均价': total_avg_price,
                    '整盘去化率': total_trans / project_total if project_total > 0 else 0,
                    '已供去化率': total_trans / total_supply if total_supply > 0 else 0,
                    '已供去化占比': '-', '已取证库存': total_supply - total_trans,
                    '整盘库存': project_total - total_trans,
                    '近3月月均销量': round(df['近3月月均销量'].sum()) if '近3月月均销量' in df.columns else 0,
                    '近3月月均销量占比': '-',
                }
                value = val_map.get(col_name, '')
                fmt = '0%' if col_name in PCT_COLS and isinstance(value, (int, float)) else None
            else:
                value = total_units_col.get(col_name, 0)
                fmt = None
            _write_cell(ws, total_row_start, col_idx, value, font=TOTAL_FONT, fmt=fmt)

        total_row_end = total_row_start + 1
        for col_idx, col_name in enumerate(all_cols, 1):
            if col_name in NON_MONTH_COLS:
                continue
            value = total_price_col.get(col_name, '')
            _write_cell(ws, total_row_end, col_idx, value, font=MONTHLY_TOTAL_FONT)

        for col_idx, col_name in enumerate(all_cols, 1):
            if col_name in NON_MONTH_COLS:
                ws.merge_cells(start_row=total_row_start, start_column=col_idx,
                               end_row=total_row_end, end_column=col_idx)

        code_col = all_cols.index('编码') + 1
        area_col = all_cols.index('户型面积') + 1
        ws.merge_cells(start_row=total_row_start, start_column=code_col,
                       end_row=total_row_end, end_column=area_col)
        mc = ws.cell(row=total_row_start, column=code_col)
        mc.value = '合计'
        mc.font = TOTAL_FONT
        mc.alignment = CENTER_WRAP

        # 板块列合并（本项目数据行 + 合计行）
        plate_col = all_cols.index('板块') + 1
        if plate:
            ws.merge_cells(start_row=data_start, start_column=plate_col,
                           end_row=total_row_end, end_column=plate_col)
            pc = ws.cell(row=data_start, column=plate_col)
            pc.value = plate
            pc.font = PLATE_FONT

        # 竞品列合并
        competitor_col = all_cols.index('竞品') + 1
        competitor_content = project_name
        if open_date:
            competitor_content += f"\n{open_date}"
        if plot_ratio:
            competitor_content += f"\n{plot_ratio}"
        ws.merge_cells(start_row=data_start, start_column=competitor_col,
                       end_row=total_row_end, end_column=competitor_col)
        cc = ws.cell(row=data_start, column=competitor_col)
        cc.value = competitor_content
        cc.font = COMPETITOR_FONT

        current_row = total_row_end + 1
        if proj_idx < len(all_results) - 1:
            current_row += 1  # 项目间空行

    # 列宽
    base_widths = {
        '板块': 8, '竞品': 10, '业态': 12, '编码': 10, '户型': 10,
        '户型面积': 10, '建面分段': 12, '整盘套数': 10, '户配': 8,
        '供应套数': 10, '成交套数': 10, '成交均价': 10,
        '整盘去化率': 10, '已供去化率': 10, '已供去化占比': 10,
        '已取证库存': 10, '整盘库存': 10,
        '近3月月均销量': 12, '近3月月均销量占比': 12,
    }
    for col_idx, col_name in enumerate(all_cols, 1):
        w = base_widths.get(col_name, 10)
        ws.column_dimensions[get_column_letter(col_idx)].width = w

    ws.freeze_panes = 'A2'
    wb.save(output_path)
    return output_path
