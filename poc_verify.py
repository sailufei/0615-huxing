"""
POC 验证脚本：从 OCR 提取的数据 → 户型合并 → 指标计算 → Excel 输出
用实际截图数据验证核心逻辑
"""
import sys
import io
# Fix Windows console encoding for Chinese output
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8', errors='replace')

import json
import re
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side, numbers
from openpyxl.utils import get_column_letter
from copy import copy

# ============================================================
# 1. 模拟 OCR 提取后的结构化数据（基于实际 EasyOCR 结果 + 人工补齐）
# ============================================================

# 供应截图数据（根据 OCR 结果整理，OCR 漏识别的数字根据合计推算补齐）
supply_raw = [
    # type: "group" = 分组行, "detail" = 户型明细行
    {"type": "group", "居室": "两室", "面积范围": "71~84", "供应套数": 0},
    {"type": "detail", "居室": "A户型-底", "面积范围": "71~73.5", "供应套数": 2, "室厅卫": "2室2厅1卫"},
    {"type": "detail", "居室": "B户型-底1", "面积范围": "76~84.5", "供应套数": 2, "室厅卫": "2室2厅1卫"},
    {"type": "group", "居室": "三室", "面积范围": "76~117", "供应套数": 336},
    {"type": "detail", "居室": "A户型", "面积范围": "75", "供应套数": 108, "室厅卫": "3室2厅1卫"},
    {"type": "detail", "居室": "B户型", "面积范围": "95", "供应套数": 183, "室厅卫": "3室2厅2卫"},
    {"type": "detail", "居室": "C1户型-底1", "面积范围": "105~109", "供应套数": 4, "室厅卫": "3室2厅2卫"},
    {"type": "detail", "居室": "C户型", "面积范围": "118", "供应套数": 41, "室厅卫": "3室2厅2卫"},
    {"type": "group", "居室": "四室", "面积范围": "116~117", "供应套数": 55},
    {"type": "detail", "居室": "C1户型", "面积范围": "118", "供应套数": 55, "室厅卫": "4室2厅2卫"},
    {"type": "group", "居室": "复式", "面积范围": "159~197", "供应套数": 30},
    {"type": "detail", "居室": "B户型-底2", "面积范围": "179~181.5", "供应套数": 2, "室厅卫": "无室厅卫"},
    {"type": "detail", "居室": "C1户型-底2", "面积范围": "159~188", "供应套数": 4, "室厅卫": "无室厅卫"},
    {"type": "detail", "居室": "C户型-底", "面积范围": "193~197", "供应套数": 24, "室厅卫": "无室厅卫"},
]

# 成交截图数据（整盘）
transaction_raw = [
    {"type": "group", "居室": "三室", "面积范围": "76~117", "成交套数": 220, "成交面积": 20273.87, "成交均价": 44097},
    {"type": "detail", "居室": "A户型", "面积范围": "75", "成交套数": 72, "成交面积": 5428.45, "成交均价": 44109, "室厅卫": "3室2厅1卫"},
    {"type": "detail", "居室": "B户型", "面积范围": "96", "成交套数": 133, "成交面积": 12863.55, "成交均价": 43879, "室厅卫": "3室2厅2卫"},
    {"type": "detail", "居室": "C户型", "面积范围": "118", "成交套数": 15, "成交面积": 1981.87, "成交均价": 45478, "室厅卫": "3室2厅2卫"},
    {"type": "group", "居室": "四室", "面积范围": "116~117", "成交套数": 26, "成交面积": 3030.90, "成交均价": 45023},
    {"type": "detail", "居室": "C1户型", "面积范围": "118", "成交套数": 26, "成交面积": 3030.90, "成交均价": 45023, "室厅卫": "4室2厅2卫"},
]

# 月度成交数据
monthly_data = {
    "2024-05": [
        {"type": "group", "居室": "两室", "面积范围": "77", "成交套数": 1, "成交面积": 77.43, "成交均价": 45039},
        {"type": "detail", "居室": "B户型-底", "面积范围": "77~81", "成交套数": 1, "成交面积": 77.43, "成交均价": 45039, "室厅卫": "2室2厅1卫"},
        {"type": "group", "居室": "三室", "面积范围": "81~97.5", "成交套数": 57, "成交面积": 4550.09, "成交均价": 51043},
        {"type": "detail", "居室": "A户型", "面积范围": "81", "成交套数": 13, "成交面积": 993.71, "成交均价": 52270, "室厅卫": "3室2厅1卫"},
        {"type": "detail", "居室": "B户型", "面积范围": "97", "成交套数": 44, "成交面积": 3555.38, "成交均价": 50699, "室厅卫": "3室2厅2卫"},
        {"type": "group", "居室": "四室", "面积范围": "112", "成交套数": 3, "成交面积": 336.06, "成交均价": 49292},
        {"type": "detail", "居室": "C户型", "面积范围": "112", "成交套数": 3, "成交面积": 336.06, "成交均价": 49292, "室厅卫": "4室2厅2卫"},
    ],
    "2024-06": [
        {"type": "group", "居室": "三室", "面积范围": "81~101", "成交套数": 38, "成交面积": 2993.75, "成交均价": 50184},
        {"type": "detail", "居室": "A户型", "面积范围": "81", "成交套数": 9, "成交面积": 678.79, "成交均价": 50794, "室厅卫": "3室2厅1卫"},
        {"type": "detail", "居室": "B户型", "面积范围": "97", "成交套数": 29, "成交面积": 2314.96, "成交均价": 50005, "室厅卫": "3室2厅2卫"},
        {"type": "group", "居室": "四室", "面积范围": "112~125", "成交套数": 4, "成交面积": 351.01, "成交均价": 54133},
        {"type": "detail", "居室": "C户型", "面积范围": "112", "成交套数": 2, "成交面积": 224.87, "成交均价": 52250, "室厅卫": "4室2厅2卫"},
        {"type": "detail", "居室": "D户型", "面积范围": "126", "成交套数": 2, "成交面积": 126.14, "成交均价": 57472, "室厅卫": "4室2厅2卫"},
    ],
    "2024-07": [
        {"type": "group", "居室": "三室", "面积范围": "81~97.5", "成交套数": 20, "成交面积": 1434.52, "成交均价": 50123},
        {"type": "detail", "居室": "A户型", "面积范围": "81", "成交套数": 7, "成交面积": 568.82, "成交均价": 50835, "室厅卫": "3室2厅1卫"},
        {"type": "detail", "居室": "B户型", "面积范围": "97", "成交套数": 13, "成交面积": 855.80, "成交均价": 49655, "室厅卫": "3室2厅2卫"},
        {"type": "group", "居室": "四室", "面积范围": "112~113", "成交套数": 3, "成交面积": 336.90, "成交均价": 51245},
        {"type": "detail", "居室": "C户型", "面积范围": "112", "成交套数": 3, "成交面积": 336.90, "成交均价": 51245, "室厅卫": "4室2厅2卫"},
    ],
}


# ============================================================
# 2. 户型变体合并逻辑（核心模块）
# ============================================================

def get_base_name(name):
    """提取户型基础名称（去掉 -顶/-底/-底1/-底2 后缀）"""
    base = re.sub(r'[-][顶底]\d*$', '', name.strip())
    return base

def is_base_unit(name):
    """判断是否是标准户型（无顶/底后缀）"""
    return not re.search(r'[-][顶底]\d*$', name.strip())

def get_bottom_number(name):
    """获取 -底 后缀的数字（-底1→1, -底→0, -底2→2）"""
    m = re.search(r'[-]底(\d*)$', name.strip())
    if m:
        if m.group(1) == '':
            return 0
        return int(m.group(1))
    return None

def merge_variants(rows, value_fields):
    """
    合并户型变体（顶/底）

    Args:
        rows: [dict] 户型明细行列表（不含 group 行）
        value_fields: [str] 需要求和的数值字段名
    Returns:
        [dict] 合并后的行列表
    """
    # 按基础名分组
    groups = {}
    for row in rows:
        base = get_base_name(row['居室'])
        if base not in groups:
            groups[base] = []
        groups[base].append(row)

    merged = []
    for base, group in groups.items():
        result = {'居室': base}

        # 求和字段
        for f in value_fields:
            result[f] = sum(r.get(f, 0) for r in group)

        # 室厅卫 / 成交均价：优先标准户型
        standard = [r for r in group if is_base_unit(r['居室'])]
        if standard:
            canonical = standard[0]
        else:
            # 无标准户型：取 -底 数字最小的
            bottom_rows = [(get_bottom_number(r['居室']), r) for r in group if get_bottom_number(r['居室']) is not None]
            if bottom_rows:
                bottom_rows.sort(key=lambda x: x[0])
                canonical = bottom_rows[0][1]
            else:
                canonical = group[0]  # fallback

        result['室厅卫'] = canonical.get('室厅卫', '')
        result['面积范围'] = canonical.get('面积范围', '')
        if '成交均价' in canonical:
            result['成交均价'] = canonical['成交均价']
        # 成交面积不聚合，取标准户型
        if '成交面积' in canonical:
            result['成交面积'] = canonical['成交面积']

        merged.append(result)

    return merged


# ============================================================
# 3. 数据解析与处理
# ============================================================

def parse_supply(rows):
    """解析供应数据：仅取明细行，合并变体"""
    details = [r for r in rows if r['type'] == 'detail']
    return merge_variants(details, ['供应套数'])

def parse_transaction(rows):
    """解析成交数据：仅取明细行，合并变体"""
    details = [r for r in rows if r['type'] == 'detail']
    return merge_variants(details, ['成交套数', '成交面积'])

def parse_monthly(rows):
    """解析月度成交数据"""
    details = [r for r in rows if r['type'] == 'detail']
    return merge_variants(details, ['成交套数', '成交面积'])


# ============================================================
# 4. 模板映射
# ============================================================

# 数字→中文
NUM_CN = {'1': '一', '2': '二', '3': '三', '4': '四', '5': '五', '6': '六'}

def map_room_type(室厅卫):
    """室厅卫 → 户型映射：3室2厅1卫 → 三室一卫"""
    if not 室厅卫 or 室厅卫 in ('复式', '无室厅卫'):
        return 室厅卫 or ''

    # 提取室和卫的数字
    shi = re.search(r'(\d+)室', 室厅卫)
    wei = re.search(r'(\d+)卫', 室厅卫)

    result = ''
    if shi:
        n = shi.group(1)
        result = NUM_CN.get(n, n) + '室'
    if wei:
        n = wei.group(1)
        result += NUM_CN.get(n, n) + '卫'

    return result if result else 室厅卫

# 建面分段
AREA_SEGMENTS = [
    (0, 65, '(0,65]'), (65, 75, '(65,75]'), (75, 85, '(75,85]'),
    (85, 95, '(85,95]'), (95, 105, '(95,105]'), (105, 115, '(105,115]'),
    (115, 125, '(115,125]'), (125, 135, '(125,135]'), (135, 150, '(135,150]'),
    (150, 165, '(150,165]'), (165, 180, '(165,180]'), (180, 200, '(180,200]'),
    (200, 220, '(200,220]'), (220, 240, '(220,240]'), (240, 260, '(240,260]'),
    (260, float('inf'), '(260,+∞)')
]

def classify_area(area_str):
    """根据面积范围字符串判断建面分段"""
    if not area_str:
        return ''
    # 提取数字，取范围中值
    nums = re.findall(r'[\d.]+', area_str)
    if not nums:
        return ''
    nums = [float(n) for n in nums]
    mid = sum(nums) / len(nums)

    for lo, hi, label in AREA_SEGMENTS:
        if lo < mid <= hi:
            return label
    return ''


# ============================================================
# 5. 主处理流程
# ============================================================

def process(project_name, plate, competitor, supply_rows, trans_rows, monthly_dict, month_labels):
    """完整处理流程"""
    supply_merged = parse_supply(supply_rows)
    trans_merged = parse_transaction(trans_rows)

    # 按户型合并供应和成交数据
    # 以供应数据为主键
    all_rows = []
    total_supply = sum(r['供应套数'] for r in supply_merged)
    total_trans = sum(r.get('成交套数', 0) for r in trans_merged)

    # 建立成交数据索引
    trans_index = {r['居室']: r for r in trans_merged}

    # 整盘套数（手动填写，此处用供应套数占位）
    整盘套数 = total_supply  # 用户手动修改

    output_rows = []
    for sup in supply_merged:
        name = sup['居室']
        trans = trans_index.get(name, {})

        sup_units = sup['供应套数']
        trans_units = trans.get('成交套数', 0)

        row = {
            '板块': plate,
            '竞品': competitor,
            '业态': '',  # 手动填写
            '户型': map_room_type(sup.get('室厅卫', '')),
            '户型面积': sup.get('面积范围', ''),
            '建面分段': classify_area(sup.get('面积范围', '')),
            '整盘套数': 整盘套数,
            '户配': sup_units / 整盘套数 if 整盘套数 else 0,
            '供应套数': sup_units,
            '成交套数': trans_units,
            '成交均价': trans.get('成交均价', ''),
            '整盘去化率': total_trans / 整盘套数 if 整盘套数 else 0,
            '已供去化率': trans_units / sup_units if sup_units else 0,
            '已供去化占比': trans_units / total_trans if total_trans else 0,
        }

        # 月度数据
        monthly_units = []
        for m_label in month_labels:
            m_data = monthly_dict.get(m_label, [])
            m_merged = parse_monthly(m_data)
            m_index = {r['居室']: r for r in m_merged}
            m_row = m_index.get(name, {})
            m_units = m_row.get('成交套数', 0)
            m_price = m_row.get('成交均价', '')
            row[f'{m_label}成交套数'] = m_units
            row[f'{m_label}成交均价'] = m_price
            monthly_units.append(m_units)

        # 近3月月均
        avg_monthly = sum(monthly_units) / len(monthly_units) if monthly_units else 0
        row['近3月月均销量'] = round(avg_monthly, 1)
        row['近3月月均销量占比'] = avg_monthly / total_trans if total_trans else 0

        output_rows.append(row)

    # 合计行
    total_row = {
        '板块': '', '竞品': '', '业态': '', '户型': '合计',
        '户型面积': '', '建面分段': '',
        '整盘套数': 整盘套数,
        '户配': 1.0,
        '供应套数': total_supply,
        '成交套数': total_trans,
        '成交均价': sum(r['成交套数'] * r['成交均价'] for r in trans_merged) / total_trans if total_trans else '',
        '整盘去化率': total_trans / 整盘套数 if 整盘套数 else 0,
        '已供去化率': total_trans / total_supply if total_supply else 0,
        '已供去化占比': 1.0,
    }
    for m_label in month_labels:
        total_row[f'{m_label}成交套数'] = sum(r[f'{m_label}成交套数'] for r in output_rows)
        # 月度均价：取整盘均价
        total_row[f'{m_label}成交均价'] = ''
    total_row['近3月月均销量'] = sum(r['近3月月均销量'] for r in output_rows)
    total_row['近3月月均销量占比'] = 1.0

    output_rows.append(total_row)

    return {
        'project_name': project_name,
        'rows': output_rows,
        'summary': {
            'total_supply': total_supply,
            'total_transaction': total_trans,
        }
    }


# ============================================================
# 6. Excel 生成
# ============================================================

def generate_excel(result, month_labels, output_path):
    """生成 Excel 文件"""
    wb = Workbook()
    ws = wb.active
    ws.title = "户型盘点"

    rows = result['rows']
    month_count = len(month_labels)

    # 列定义
    base_headers = ['板块', '竞品', '业态', '户型', '户型面积', '建面分段',
                    '整盘套数', '户配', '供应套数', '成交套数', '成交均价',
                    '整盘去化率', '已供去化率', '已供去化占比']
    month_cols = []
    for ml in month_labels:
        month_cols.append(f'{ml}\n成交套数')
        month_cols.append(f'{ml}\n成交均价')
    trailing = ['近3月月均销量', '近3月月均销量占比']

    all_headers = base_headers + month_cols + trailing

    # 手动填写的列索引（黄色标记）
    MANUAL_COLS = {'板块', '竞品', '业态', '整盘套数'}

    # 样式
    header_font = Font(name='微软雅黑', bold=True, size=10)
    header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
    header_font_white = Font(name='微软雅黑', bold=True, size=10, color='FFFFFF')
    manual_fill = PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid')  # 黄色
    thin_border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )
    center_align = Alignment(horizontal='center', vertical='center', wrap_text=True)

    # 写入表头
    for col_idx, header in enumerate(all_headers, 1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = header_font_white
        cell.fill = header_fill
        cell.alignment = center_align
        cell.border = thin_border

    # 写入数据
    for row_idx, row_data in enumerate(rows, 2):
        for col_idx, header in enumerate(all_headers, 1):
            value = row_data.get(header.replace('\n成交套数', '成交套数').replace('\n成交均价', '成交均价'), '')

            cell = ws.cell(row=row_idx, column=col_idx)
            cell.border = thin_border
            cell.alignment = center_align

            # 百分比格式
            if header in ('户配', '整盘去化率', '已供去化率', '已供去化占比', '近3月月均销量占比'):
                if isinstance(value, (int, float)):
                    cell.value = value
                    cell.number_format = '0%'
                else:
                    cell.value = value
            # 均价格式
            elif '均价' in header:
                if isinstance(value, (int, float)):
                    cell.value = value
                    cell.number_format = '#,##0'
                else:
                    cell.value = value
            else:
                cell.value = value

            # 黄色标记手动填写列
            base_col = header.replace('\n成交套数', '').replace('\n成交均价', '')
            if base_col in MANUAL_COLS:
                cell.fill = manual_fill

            # 合计行加粗
            if row_data.get('户型') == '合计':
                cell.font = Font(name='微软雅黑', bold=True, size=10)

    # 列宽
    col_widths = {
        '板块': 8, '竞品': 10, '业态': 12, '户型': 10, '户型面积': 12,
        '建面分段': 12, '整盘套数': 10, '户配': 8, '供应套数': 10,
        '成交套数': 10, '成交均价': 10, '整盘去化率': 10, '已供去化率': 10,
        '已供去化占比': 10,
    }
    for col_idx, header in enumerate(all_headers, 1):
        base = header.replace('\n成交套数', '').replace('\n成交均价', '')
        width = col_widths.get(base, 12)
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    # 冻结首行
    ws.freeze_panes = 'A2'

    wb.save(output_path)
    return output_path


# ============================================================
# 7. 终端预览
# ============================================================

def print_preview(result, month_labels):
    """终端打印预览"""
    rows = result['rows']
    print(f"\n{'='*80}")
    print(f"  项目: {result['project_name']}")
    print(f"  整盘供应: {result['summary']['total_supply']} 套")
    print(f"  整盘成交: {result['summary']['total_transaction']} 套")
    print(f"{'='*80}\n")

    # 简化预览
    preview_cols = ['户型', '户型面积', '建面分段', '供应套数', '成交套数', '成交均价', '已供去化占比']
    for ml in month_labels:
        preview_cols.append(f'{ml}成交套数')

    # 表头
    header_line = ' | '.join(f'{c:^10}' for c in preview_cols)
    print(header_line)
    print('-' * len(header_line))

    for row in rows:
        values = []
        for c in preview_cols:
            v = row.get(c, '')
            if isinstance(v, float):
                if '占比' in c or '率' in c:
                    v = f'{v:.0%}'
                elif '均价' in c:
                    v = f'{v:,.0f}'
                else:
                    v = f'{v}'
            values.append(f'{str(v):^10}')
        print(' | '.join(values))


# ============================================================
# 8. 执行验证
# ============================================================

if __name__ == '__main__':
    month_labels = ['2024-05', '2024-06', '2024-07']

    result = process(
        project_name='华曦府金安（POC验证）',
        plate='',  # 手动填
        competitor='',  # 手动填
        supply_rows=supply_raw,
        trans_rows=transaction_raw,
        monthly_dict=monthly_data,
        month_labels=month_labels,
    )

    # 打印预览
    print_preview(result, month_labels)

    # 生成 Excel
    output_path = r'd:\guojiabao008\Desktop\Cursor workspace\POC_验证输出.xlsx'
    generate_excel(result, month_labels, output_path)

    print(f"\n[OK] Excel saved: {output_path}")
    print(f"   Manual-fill columns (yellow): plate, competitor, property_type, total_units")
    print(f"   Please open the Excel file to verify format and data.")

    # 详细数据
    print(f"\n{'='*80}")
    print("  详细行数据 (JSON):")
    print(f"{'='*80}")
    for i, row in enumerate(result['rows']):
        print(f"\n--- 行 {i+1}: {row['户型']} ---")
        for k, v in row.items():
            if isinstance(v, float):
                print(f"  {k}: {v:.4f}" if v < 1 else f"  {k}: {v:,.2f}")
            else:
                print(f"  {k}: {v}")
